"""One-click Excel workbook export of the cohort comparison results.

Sheets (in reading order):
    Summary       - run scope, windows, headline numbers, flag legend
    Deltas        - volumes, absolute / % change, contribution share
    Effect sizes  - Cohen's d, Welch t, p, BH q, significance
    Driver flags  - weather deltas, driver-day share, calendar-mix warning
    Narrative     - the auto-written case-study text

Written with openpyxl so the sheets carry number formats, frozen panes,
auto-filters and conditional colour scales rather than raw strings.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

KEYS = ["borough", "complaint_type"]

DELTA_COLUMNS = KEYS + [
    "baseline_days", "post_days", "baseline_volume", "post_volume",
    "baseline_daily_avg", "post_daily_avg", "baseline_volume_scaled",
    "abs_change", "pct_change", "contribution_pct",
]
EFFECT_COLUMNS = KEYS + [
    "effect_size_d", "t_statistic", "p_value", "q_value", "is_significant",
]
DRIVER_COLUMNS = KEYS + [
    "temp_max_f_delta", "precip_in_delta", "baseline_driver_day_pct",
    "post_driver_day_pct", "driver_day_pct_delta", "calendar_mix_warning",
]

NUMBER_FORMATS = {
    "baseline_volume": "#,##0", "post_volume": "#,##0",
    "baseline_volume_scaled": "#,##0", "abs_change": "+#,##0;-#,##0;0",
    "baseline_daily_avg": "#,##0.0", "post_daily_avg": "#,##0.0",
    "pct_change": "+0.0\\%;-0.0\\%;0.0\\%",
    "contribution_pct": "+0.0\\%;-0.0\\%;0.0\\%",
    "driver_day_pct_delta": "+0.0\\ \"pp\";-0.0\\ \"pp\";0.0\\ \"pp\"",
    "baseline_driver_day_pct": "0.0\\%", "post_driver_day_pct": "0.0\\%",
    "effect_size_d": "+0.00;-0.00;0.00", "t_statistic": "0.00",
    "p_value": "0.0000", "q_value": "0.0000",
    "temp_max_f_delta": "+0.0;-0.0;0.0", "precip_in_delta": "+0.00;-0.00;0.00",
}

HEADER_FILL = "1F3864"


def _subset(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    return df.reindex(columns=[c for c in columns if c in df.columns]).copy()


def _style_sheet(ws, df: pd.DataFrame, heat_columns: tuple[str, ...] = ()) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "C2"
    for idx, name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = max(len(str(name)) + 2, 12)
        if name in ("complaint_type",):
            width = 34
        elif name == "borough":
            width = 18
        ws.column_dimensions[letter].width = width
        fmt = NUMBER_FORMATS.get(name)
        if fmt:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=idx).number_format = fmt
        if name in heat_columns and len(df):
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{len(df) + 1}",
                ColorScaleRule(start_type="min", start_color="F4B183",
                               mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                               end_type="max", end_color="9BC2E6"))
    if len(df):
        ws.auto_filter.ref = ws.dimensions


def _write_summary(ws, df: pd.DataFrame, meta: dict) -> None:
    from openpyxl.styles import Alignment, Font

    rows = [
        ("NYC 311 — cohort comparison export", ""),
        ("Generated (UTC)", f"{datetime.now(timezone.utc):%Y-%m-%d %H:%M}"),
        ("Scope", meta.get("scope", "All boroughs, all complaint categories")),
        ("Windows", meta.get("window_label", "see config/config.yaml")),
        ("Cohorts", len(df)),
        ("Net change (requests)", float(df["abs_change"].sum(skipna=True)) if len(df) else 0.0),
        ("Significant cohorts (BH q < alpha)",
         int(df.get("is_significant", pd.Series(dtype=bool)).fillna(False).sum())),
        ("Alpha", meta.get("alpha", 0.01)),
        ("Calendar-mix warnings",
         int(df.get("calendar_mix_warning", pd.Series(dtype=bool)).fillna(False).sum())),
        ("", ""),
        ("Sheet guide", ""),
        ("Deltas", "Volumes and change: length-normalised baseline, abs/% change, contribution share"),
        ("Effect sizes", "Cohen's d, Welch t-statistic, p-value, Benjamini-Hochberg q, significance"),
        ("Driver flags", "Weather deltas, driver-day share shift, calendar-mix warning"),
        ("Narrative", "Auto-written case-study summary of drivers and confounder eliminations"),
    ]
    for r, (key, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=key).font = Font(bold=(r == 1 or key == "Sheet guide"))
        ws.cell(row=r, column=2, value=value).alignment = Alignment(wrap_text=True)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 96


def _write_narrative(ws, text: str) -> None:
    from openpyxl.styles import Alignment, Font

    ws.column_dimensions["A"].width = 120
    row = 1
    for line in (text or "").splitlines():
        stripped = line.strip()
        if not stripped:
            row += 1
            continue
        cell = ws.cell(row=row, column=1, value=stripped.lstrip("# ").replace("**", ""))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if stripped.startswith("#"):
            cell.font = Font(bold=True, size=12 if stripped.startswith("##") else 14)
        row += 1


def export_excel(df: pd.DataFrame, path: str | Path, *, narrative: str = "",
                 meta: dict | None = None) -> Path:
    """Write the multi-sheet workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    meta = meta or {}
    wb = Workbook()

    _write_summary(wb.active, df, meta)
    wb.active.title = "Summary"

    sheets = [
        ("Deltas", _subset(df, DELTA_COLUMNS), ("abs_change", "pct_change")),
        ("Effect sizes", _subset(df, EFFECT_COLUMNS), ("effect_size_d",)),
        ("Driver flags", _subset(df, DRIVER_COLUMNS), ("driver_day_pct_delta",)),
    ]
    for name, frame, heat in sheets:
        ws = wb.create_sheet(name)
        for row in dataframe_to_rows(frame, index=False, header=True):
            ws.append(["" if pd.isna(v) else v for v in row])
        _style_sheet(ws, frame, heat)

    _write_narrative(wb.create_sheet("Narrative"), narrative)
    wb.save(path)
    return path
