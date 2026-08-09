"""Tests for the scheduled/filtered/emailed export layer."""
from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from src.reporting import cohort_export as ce
from src.reporting import excel_export, narrative, recent_week
from src.reporting.emailer import build_message, parse_recipients, send_exports
from src.reporting.filters import CohortFilter, recompute_contribution


# ------------------------------------------------------------------ fixtures
@pytest.fixture()
def enriched_daily() -> pd.DataFrame:
    days = pd.date_range("2024-05-01", "2024-06-30", freq="D")
    rows = []
    rng = np.random.default_rng(7)
    for borough in ("BROOKLYN", "QUEENS"):
        for ctype in ("Noise - Residential", "Heat/Hot Water"):
            base = 100 if borough == "BROOKLYN" else 60
            for day in days:
                bump = 40 if (borough == "BROOKLYN" and day >= pd.Timestamp("2024-06-20")) else 0
                rows.append({
                    "date_day": day, "borough": borough, "complaint_type": ctype,
                    "volume": base + bump + rng.integers(0, 5),
                    "temp_max_f": 78 + rng.integers(0, 6),
                    "precip_in": 0.0,
                    "is_hot_day": False, "is_freezing_day": False,
                    "is_heavy_rain_day": False, "is_snow_day": False,
                    "is_high_wind_day": False, "is_holiday": False,
                    "is_weekend": day.dayofweek >= 5,
                    "is_non_business_day": day.dayofweek >= 5,
                })
    return pd.DataFrame(rows)


# ------------------------------------------------------------------- filters
def test_filter_parses_all_tokens_and_strings():
    f = CohortFilter.from_dict({"boroughs": "BROOKLYN, (All)", "complaint_type": None})
    assert f.boroughs == ["BROOKLYN"] and f.complaint_types == []
    assert not f.is_empty and f.slug() == "brooklyn"
    assert CohortFilter().is_empty


def test_filter_applies_to_daily_and_dates(enriched_daily):
    f = CohortFilter(boroughs=["queens"], date_from="2024-06-01")
    out = f.apply(enriched_daily)
    assert set(out["borough"]) == {"QUEENS"}
    assert pd.to_datetime(out["date_day"]).min() == pd.Timestamp("2024-06-01")


def test_filter_json_roundtrip(tmp_path):
    path = tmp_path / "pbi_filters.json"
    path.write_text('{"boroughs":["BRONX"],"complaint_types":["Noise - Residential"],'
                    '"source":"Power BI cohort page"}')
    f = CohortFilter.from_json(path)
    assert f.boroughs == ["BRONX"] and f.source == "Power BI cohort page"


def test_contribution_rebases_to_selection():
    df = pd.DataFrame({"abs_change": [30.0, 10.0], "contribution_pct": [3.0, 1.0]})
    out = recompute_contribution(df)
    assert out["contribution_pct"].sum() == pytest.approx(100.0)


# -------------------------------------------------------------- recent week
def test_recent_week_windows_respect_lag(enriched_daily):
    w = recent_week.recent_week_windows(enriched_daily, lag_days=1)
    assert str(w.post_end.date()) == "2024-06-29"
    assert (w.post_end - w.post_start).days == 6
    assert (w.baseline_end - w.baseline_start).days == 27
    assert w.baseline_end < w.post_start


def test_recent_week_cohorts_detect_the_bump(enriched_daily):
    comparison, tagged, windows = recent_week.build_recent_week_cohorts(enriched_daily)
    assert set(tagged["cohort_period"]) == {"Baseline", "Post-change"}
    brooklyn = comparison[(comparison.borough == "BROOKLYN")
                          & (comparison.complaint_type == "Noise - Residential")].iloc[0]
    queens = comparison[(comparison.borough == "QUEENS")
                        & (comparison.complaint_type == "Heat/Hot Water")].iloc[0]
    assert brooklyn.abs_change > 200 and brooklyn["pct_change"] > 20
    assert abs(queens["pct_change"]) < 5
    assert comparison["contribution_pct"].sum() == pytest.approx(100.0)
    assert windows.label().startswith("post 2024-06-23")


# --------------------------------------------------------------- narrative
def test_narrative_names_top_driver_and_confounders(enriched_daily):
    comparison, tagged, windows = recent_week.build_recent_week_cohorts(enriched_daily)
    df = ce.build_export_frame(comparison, tagged)
    text = narrative.build_narrative(df, tagged, scope="All boroughs",
                                     window_label=windows.label())
    assert "BROOKLYN / Noise - Residential" in text
    assert "Confounders tested and eliminated" in text
    assert "ELIMINATED" in text and "Benjamini-Hochberg" in text
    assert "Conclusion" in text and "Recommended next steps" in text


def test_narrative_handles_empty_selection():
    text = narrative.build_narrative(pd.DataFrame(), None, scope="BRONX")
    assert "No cohorts matched" in text


def test_confounder_checks_flag_weather(enriched_daily):
    comparison, tagged, _ = recent_week.build_recent_week_cohorts(enriched_daily)
    df = ce.build_export_frame(comparison, tagged)
    df.loc[0, "driver_day_pct_delta"] = 45.0
    checks = {c["name"]: c for c in narrative.confounder_checks(df, tagged)}
    assert checks["Weather (heat, rain, snow, wind)"]["eliminated"] is False


# ------------------------------------------------------------------- excel
def test_excel_has_the_three_result_sheets(enriched_daily, tmp_path):
    from openpyxl import load_workbook

    comparison, tagged, _ = recent_week.build_recent_week_cohorts(enriched_daily)
    df = ce.build_export_frame(comparison, tagged)
    path = excel_export.export_excel(df, tmp_path / "cohorts.xlsx",
                                     narrative="# Title\n\nBody line.")
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Deltas", "Effect sizes", "Driver flags",
                             "Narrative"]
    assert wb["Deltas"].max_row == len(df) + 1
    assert [c.value for c in wb["Effect sizes"][1]][:3] == [
        "borough", "complaint_type", "effect_size_d"]
    assert "calendar_mix_warning" in [c.value for c in wb["Driver flags"][1]]


# ------------------------------------------------------------------- email
def test_parse_recipients_dedupes_and_validates():
    assert parse_recipients("a@x.com, b@x.com; a@x.com  bad") == ["a@x.com", "b@x.com"]
    assert parse_recipients(None) == []


def test_build_message_attaches_files(tmp_path):
    csv = tmp_path / "r.csv"
    csv.write_text("a,b\n1,2\n")
    msg = build_message([csv], ["ops@example.com"], sender="bot@example.com",
                        subject="S", body_text="hello", body_markdown="# H\n\nhi")
    names = [p.get_filename() for p in msg.iter_attachments() if p.get_filename()]
    assert names == ["r.csv"]
    assert msg["To"] == "ops@example.com"


def test_send_exports_dry_run_requires_recipients(tmp_path):
    with pytest.raises(ValueError):
        send_exports([], [], subject="S", body_text="b", dry_run=True)
    out = send_exports([], ["ops@example.com"], subject="S", body_text="b", dry_run=True)
    assert out["dry_run"] is True and out["to"] == ["ops@example.com"]


# ------------------------------------------------------- end-to-end pipeline
def test_run_recent_week_filtered_all_formats(enriched_daily, tmp_path):
    marts = tmp_path / "marts"
    marts.mkdir()
    enriched_daily.to_parquet(marts / "mart_daily_volume_enriched_by_type.parquet")
    outcome = ce.run(out_dir=tmp_path / "out", fmt="all", from_parquet=marts,
                     filters=CohortFilter(boroughs=["BROOKLYN"], source="test"),
                     recent_week=True, stem="nightly",
                     email=True, email_to="ops@example.com", email_dry_run=True)
    files = outcome["files"]
    assert files["csv"].exists() and files["xlsx"].exists()
    assert files["pdf"].read_bytes().startswith(b"%PDF")
    assert files["narrative"].read_text().startswith("# Cohort comparison")
    assert "brooklyn" in files["csv"].name
    exported = pd.read_csv(files["csv"])
    assert set(exported["borough"]) == {"BROOKLYN"}
    assert outcome["window"]["post_end"] == "2024-06-29"
    assert outcome["email"]["dry_run"] and outcome["email"]["to"] == ["ops@example.com"]
    assert sorted(outcome["email"]["attachments"]) == [
        "nightly_brooklyn.csv", "nightly_brooklyn.pdf", "nightly_brooklyn.xlsx"]


def test_nightly_job_logs_a_run(enriched_daily, tmp_path):
    from src.orchestration import nightly_export

    marts = tmp_path / "marts"
    marts.mkdir()
    enriched_daily.to_parquet(marts / "mart_daily_volume_enriched_by_type.parquet")
    result = nightly_export.run(out_dir=tmp_path / "nightly", from_parquet=marts,
                                email=True, email_to="ops@example.com",
                                email_dry_run=True)
    assert (tmp_path / "nightly" / "nightly_runs.jsonl").exists()
    assert result["cohorts"] > 0 and result["window"]["post_start"] == "2024-06-23"
